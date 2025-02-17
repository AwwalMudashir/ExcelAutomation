import openpyxl as xl
from openpyxl.chart import BarChart, Reference
import os


def process_workbook(filename):
  file_path = os.path.join(os.path.dirname(__file__), filename)
  wb = xl.load_workbook(file_path)
  # The above is loading a workbook and returns a workbook object
  sheet = wb['Sheet1'] # also returns a sheet object



  for row in range(2, sheet.max_row + 1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price

  # we use Reference to get the range of values
  values = Reference(sheet, min_row=2, max_row=sheet.max_row, min_col=4, max_col=4)

  chart = BarChart()
  chart.add_data(values)
  sheet.add_chart(chart, 'e2')

  wb.save(filename)





# Comments

# The above function takes in a filename and processes the workbook by loading the workbook, getting the sheet, iterating through the rows, and then saving the workbook

# Methods of Accessing a Cell

# Method 1:
# cell = sheet['a1'] # passing the name of the cell

# Method 2:
# cell = sheet.cell(1,1) # the arguements passed are the numbers for the row and column

# print(cell.value)
# print(sheet.max_row) # returns the number of rows in the sheet

process_workbook('transactions.xlsx')